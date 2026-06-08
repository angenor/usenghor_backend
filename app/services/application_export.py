"""
Service d'export des candidatures
=================================

Génère une archive ZIP contenant un sous-dossier par candidat. Chaque dossier
regroupe l'ensemble des documents soumis (CV, diplômes, pièces justificatives)
ainsi qu'un fichier Excel récapitulant ses informations.
"""

import csv
import io
import os
import re
import unicodedata
import zipfile
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings

from app.models.application import (
    Application,
    ApplicationDocument,
    EmploymentStatus,
    ExperienceDuration,
    MaritalStatus,
    SubmittedApplicationStatus,
)
from app.models.base import Salutation
from app.services.application_service import ApplicationService
from app.services.media_service import MediaService

# Libellés français pour l'export
SALUTATION_LABELS: dict[Salutation, str] = {
    Salutation.Mr: "M.",
    Salutation.Mrs: "Mme",
    Salutation.Dr: "Dr",
    Salutation.Pr: "Pr",
}

STATUS_LABELS: dict[SubmittedApplicationStatus, str] = {
    SubmittedApplicationStatus.SUBMITTED: "Soumise",
    SubmittedApplicationStatus.UNDER_REVIEW: "En évaluation",
    SubmittedApplicationStatus.ACCEPTED: "Acceptée",
    SubmittedApplicationStatus.REJECTED: "Rejetée",
    SubmittedApplicationStatus.WAITLISTED: "Liste d'attente",
    SubmittedApplicationStatus.INCOMPLETE: "Incomplète",
}

MARITAL_STATUS_LABELS: dict[MaritalStatus, str] = {
    MaritalStatus.SINGLE: "Célibataire",
    MaritalStatus.MARRIED: "Marié(e)",
    MaritalStatus.DIVORCED: "Divorcé(e)",
    MaritalStatus.WIDOWED: "Veuf/Veuve",
    MaritalStatus.OTHER: "Autre",
}

EMPLOYMENT_STATUS_LABELS: dict[EmploymentStatus, str] = {
    EmploymentStatus.STUDENT: "Étudiant(e)",
    EmploymentStatus.TEACHER: "Enseignant(e)",
    EmploymentStatus.CIVIL_SERVANT: "Fonctionnaire",
    EmploymentStatus.PRIVATE_EMPLOYEE: "Employé(e) du privé",
    EmploymentStatus.NGO_EMPLOYEE: "Employé(e) ONG",
    EmploymentStatus.UNEMPLOYED: "Sans emploi",
    EmploymentStatus.OTHER: "Autre",
}

EXPERIENCE_DURATION_LABELS: dict[ExperienceDuration, str] = {
    ExperienceDuration.LESS_THAN_1_YEAR: "Moins d'1 an",
    ExperienceDuration.BETWEEN_1_3_YEARS: "1 à 3 ans",
    ExperienceDuration.BETWEEN_3_5_YEARS: "3 à 5 ans",
    ExperienceDuration.BETWEEN_5_10_YEARS: "5 à 10 ans",
    ExperienceDuration.MORE_THAN_10_YEARS: "Plus de 10 ans",
}

_SECTION_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")


def _slugify(value: str, *, fallback: str = "candidat") -> str:
    """Translittère et nettoie une chaîne pour un nom de fichier/dossier sûr.

    Supprime les accents (problèmes d'encodage SSH/Docker) et ne conserve que
    des caractères ASCII alphanumériques, tirets et underscores.
    """
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_str = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", ascii_str).strip("_")
    return cleaned or fallback


def _format_value(value: object) -> str:
    """Formate une valeur de candidature pour affichage dans l'Excel."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, Decimal):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    if isinstance(value, Salutation):
        return SALUTATION_LABELS.get(value, value.value)
    if isinstance(value, SubmittedApplicationStatus):
        return STATUS_LABELS.get(value, value.value)
    if isinstance(value, MaritalStatus):
        return MARITAL_STATUS_LABELS.get(value, value.value)
    if isinstance(value, EmploymentStatus):
        return EMPLOYMENT_STATUS_LABELS.get(value, value.value)
    if isinstance(value, ExperienceDuration):
        return EXPERIENCE_DURATION_LABELS.get(value, value.value)
    return str(value)


class ApplicationExportService:
    """Construit l'archive ZIP d'export des candidatures."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.application_service = ApplicationService(db)
        self.media_service = MediaService(db)

    async def build_zip(
        self,
        *,
        search: str | None = None,
        call_id: str | None = None,
        status: SubmittedApplicationStatus | None = None,
        program_id: str | None = None,
        ids: list[str] | None = None,
    ) -> tuple[io.BytesIO, str]:
        """Construit l'archive ZIP des candidatures filtrées.

        Args:
            search, call_id, status, program_id: Mêmes filtres que la liste.
            ids: Restreint l'export à ces identifiants de candidature.

        Returns:
            Tuple (buffer ZIP positionné au début, nom de fichier suggéré).
        """
        query = await self.application_service.get_applications(
            search=search,
            call_id=call_id,
            status=status,
            program_id=program_id,
        )
        query = query.options(selectinload(Application.call))
        result = await self.db.execute(query)
        applications = list(result.scalars().unique().all())

        if ids:
            id_set = set(ids)
            applications = [app for app in applications if app.id in id_set]

        zip_buffer = io.BytesIO()
        used_folders: set[str] = set()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for application in applications:
                folder = self._unique_folder_name(application, used_folders)

                # Fichier Excel des informations
                zip_file.writestr(
                    f"{folder}/informations.xlsx",
                    self._build_xlsx(application),
                )

                # Documents soumis (CV, diplômes, justificatifs...)
                await self._add_documents(zip_file, folder, application)

        zip_buffer.seek(0)
        filename = f"candidatures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return zip_buffer, filename

    async def build_csv(
        self,
        *,
        search: str | None = None,
        call_id: str | None = None,
        status: SubmittedApplicationStatus | None = None,
        program_id: str | None = None,
        ids: list[str] | None = None,
    ) -> tuple[io.BytesIO, str]:
        """Construit un unique fichier CSV récapitulatif des candidatures.

        Une ligne par candidat avec ses informations principales, suivies de
        colonnes « Document N » / « Lien document N » pointant (en URL absolue)
        vers les fichiers téléversés (CV, diplômes, justificatifs...).

        Args:
            search, call_id, status, program_id: Mêmes filtres que la liste.
            ids: Restreint l'export à ces identifiants de candidature.

        Returns:
            Tuple (buffer CSV positionné au début, nom de fichier suggéré).
        """
        query = await self.application_service.get_applications(
            search=search,
            call_id=call_id,
            status=status,
            program_id=program_id,
        )
        query = query.options(selectinload(Application.call))
        result = await self.db.execute(query)
        applications = list(result.scalars().unique().all())

        if ids:
            id_set = set(ids)
            applications = [app for app in applications if app.id in id_set]

        # Documents triés par date d'ajout pour un ordre de colonnes stable.
        docs_by_app = {
            app.id: sorted(app.documents, key=lambda d: d.created_at)
            for app in applications
        }
        max_docs = max((len(docs) for docs in docs_by_app.values()), default=0)
        base = settings.frontend_url.rstrip("/")

        header = [
            "Référence",
            "Appel à candidature",
            "Statut",
            "Date de soumission",
            "Score d'évaluation",
            "Civilité",
            "Nom",
            "Prénom",
            "Date de naissance",
            "Ville de naissance",
            "Situation matrimoniale",
            "Email",
            "Téléphone",
            "WhatsApp",
            "Adresse",
            "Ville",
            "Code postal",
            "Statut professionnel",
            "Expérience professionnelle",
            "Durée d'expérience",
            "Poste actuel",
            "Employeur",
            "Plus haut niveau de diplôme",
            "Intitulé du diplôme",
        ]
        for index in range(1, max_docs + 1):
            header.extend([f"Document {index}", f"Lien document {index}"])

        output = io.StringIO()
        writer = csv.writer(output, delimiter=";")
        writer.writerow(header)

        for application in applications:
            row = [
                _format_value(application.reference_number),
                application.call.title if application.call else "Candidature spontanée",
                _format_value(application.status),
                _format_value(application.submitted_at),
                _format_value(application.review_score),
                _format_value(application.salutation),
                _format_value(application.last_name),
                _format_value(application.first_name),
                _format_value(application.birth_date),
                _format_value(application.birth_city),
                _format_value(application.marital_status),
                _format_value(application.email),
                _format_value(application.phone),
                _format_value(application.phone_whatsapp),
                _format_value(application.address),
                _format_value(application.city),
                _format_value(application.postal_code),
                _format_value(application.employment_status),
                _format_value(application.has_work_experience),
                _format_value(application.experience_duration),
                _format_value(application.job_title or application.current_job),
                _format_value(application.employer_name),
                _format_value(application.highest_degree_level),
                _format_value(application.highest_degree_title),
            ]
            for document in docs_by_app[application.id]:
                row.append(_format_value(document.document_name) or "Document")
                if document.media_external_id and base:
                    row.append(
                        f"{base}/api/public/media/"
                        f"{document.media_external_id}/download?download=1"
                    )
                else:
                    row.append("")
            writer.writerow(row)

        # Encodage UTF-8 avec BOM pour une ouverture correcte des accents
        # dans Excel (séparateur point-virgule, locale francophone).
        csv_buffer = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        csv_buffer.seek(0)
        filename = f"candidatures_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return csv_buffer, filename

    def _unique_folder_name(
        self, application: Application, used_folders: set[str]
    ) -> str:
        """Génère un nom de dossier lisible et unique pour un candidat."""
        base = _slugify(
            f"{application.last_name}_{application.first_name}_{application.reference_number}"
        )
        folder = base
        counter = 2
        while folder in used_folders:
            folder = f"{base}_{counter}"
            counter += 1
        used_folders.add(folder)
        return folder

    async def _add_documents(
        self,
        zip_file: zipfile.ZipFile,
        folder: str,
        application: Application,
    ) -> None:
        """Ajoute tous les documents soumis du candidat dans son dossier.

        Les documents sans fichier associé ou introuvables sur le disque sont
        ignorés silencieusement afin de ne pas faire échouer l'export complet.
        """
        used_names: set[str] = set()
        for document in application.documents:
            if not document.media_external_id:
                continue
            try:
                file_path, media_name, _mime = await self.media_service.get_download_info(
                    document.media_external_id
                )
                data = file_path.read_bytes()
            except Exception:
                continue

            entry_name = self._unique_document_name(document, media_name, used_names)
            zip_file.writestr(f"{folder}/{entry_name}", data)

    @staticmethod
    def _unique_document_name(
        document: ApplicationDocument,
        media_name: str | None,
        used_names: set[str],
    ) -> str:
        """Construit un nom de fichier lisible et unique pour un document."""
        # Extension issue du nom du média (ex: .pdf, .docx)
        _, ext = os.path.splitext(media_name or "")
        ext = ext if re.fullmatch(r"\.[A-Za-z0-9]{1,8}", ext or "") else ""

        base = _slugify(document.document_name or "document", fallback="document")
        name = f"{base}{ext}"
        counter = 2
        while name in used_names:
            name = f"{base}_{counter}{ext}"
            counter += 1
        used_names.add(name)
        return name

    def _build_xlsx(self, application: Application) -> bytes:
        """Génère le classeur Excel récapitulatif d'une candidature."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Candidature"

        label_font = Font(bold=True)
        section_font = Font(bold=True, size=12, color="FFFFFF")
        wrap = Alignment(wrap_text=True, vertical="top")

        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["B"].width = 60

        row = 1

        def add_section(title: str) -> None:
            nonlocal row
            cell = sheet.cell(row=row, column=1, value=title)
            cell.font = section_font
            cell.fill = _SECTION_FILL
            sheet.cell(row=row, column=2).fill = _SECTION_FILL
            row += 1

        def add_field(label: str, value: object) -> None:
            nonlocal row
            label_cell = sheet.cell(row=row, column=1, value=label)
            label_cell.font = label_font
            label_cell.alignment = wrap
            value_cell = sheet.cell(row=row, column=2, value=_format_value(value))
            value_cell.alignment = wrap
            row += 1

        add_section("Référence")
        add_field("Numéro de référence", application.reference_number)
        add_field(
            "Appel à candidature",
            application.call.title if application.call else "Candidature spontanée",
        )
        add_field("Statut", application.status)
        add_field("Date de soumission", application.submitted_at)
        add_field("Score d'évaluation", application.review_score)

        add_section("Informations personnelles")
        add_field("Civilité", application.salutation)
        add_field("Nom", application.last_name)
        add_field("Prénom", application.first_name)
        add_field("Date de naissance", application.birth_date)
        add_field("Ville de naissance", application.birth_city)
        add_field("Situation matrimoniale", application.marital_status)

        add_section("Coordonnées")
        add_field("Email", application.email)
        add_field("Téléphone", application.phone)
        add_field("WhatsApp", application.phone_whatsapp)
        add_field("Adresse", application.address)
        add_field("Ville", application.city)
        add_field("Code postal", application.postal_code)

        add_section("Situation professionnelle")
        add_field("Statut professionnel", application.employment_status)
        add_field("Précision (autre)", application.employment_status_other)
        add_field("Expérience professionnelle", application.has_work_experience)
        add_field("Poste actuel", application.current_job)
        add_field("Intitulé du poste", application.job_title)
        add_field("Employeur", application.employer_name)
        add_field("Adresse employeur", application.employer_address)
        add_field("Ville employeur", application.employer_city)
        add_field("Téléphone employeur", application.employer_phone)
        add_field("Email employeur", application.employer_email)
        add_field("Durée d'expérience", application.experience_duration)

        add_section("Formation académique")
        add_field("Plus haut niveau de diplôme", application.highest_degree_level)
        add_field("Intitulé du diplôme", application.highest_degree_title)
        add_field("Date d'obtention", application.degree_date)
        add_field("Lieu d'obtention", application.degree_location)

        if application.degrees:
            add_section("Diplômes")
            for index, degree in enumerate(application.degrees, start=1):
                add_field(
                    f"Diplôme {index}",
                    " — ".join(
                        part
                        for part in [
                            degree.title,
                            str(degree.year) if degree.year else None,
                            degree.institution,
                            degree.specialization,
                            degree.honors,
                        ]
                        if part
                    ),
                )

        if application.motivation_text:
            add_section("Motivation")
            add_field("Texte de motivation", application.motivation_text)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
